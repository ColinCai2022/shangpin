import os

from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_from_directory,
    url_for,
)

from auth import login_required
from models import Product, db

product_bp = Blueprint("products", __name__, url_prefix="/products")


@product_bp.route("/uploads/<path:filename>")
def uploaded_file(filename):
    upload_folder = current_app.config["UPLOAD_FOLDER"]
    return send_from_directory(upload_folder, filename)


@product_bp.route("/", methods=["GET"])
@login_required()
def list_products():
    # 简单查询：按型号或 SKU 搜索
    keyword = request.args.get("q", "").strip()
    page = int(request.args.get("page", 1))
    per_page = 20

    query = Product.query
    if keyword:
        like = f"%{keyword}%"
        query = query.filter(
            (Product.model.ilike(like)) | (Product.sku.ilike(like))
        )

    pagination = query.order_by(Product.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return render_template(
        "product_list.html",
        pagination=pagination,
        products=pagination.items,
        keyword=keyword,
    )


@product_bp.route("/create", methods=["GET", "POST"])
@login_required("admin")
def create_product():
    if request.method == "POST":
        form = request.form
        files = request.files
        product = Product(
            model=form.get("model", "").strip(),
            sku=form.get("sku", "").strip(),
            package_size=form.get("package_size", "").strip(),
            outer_size=form.get("outer_size", "").strip(),
            inner_size=form.get("inner_size", "").strip(),
            gross_weight=form.get("gross_weight", "").strip(),
            net_weight=form.get("net_weight", "").strip(),
            temperature_range=form.get("temperature_range", "").strip(),
            power=form.get("power", "").strip(),
            power_consumption=form.get("power_consumption", "").strip(),
            compressor_brand=form.get("compressor_brand", "").strip(),
            cooling_pipe_material=form.get("cooling_pipe_material", "").strip(),
            energy_level=form.get("energy_level", "").strip(),
        )

        save_uploaded_images(product, files)

        if not product.model:
            flash("型号为必填项", "danger")
            return render_template("product_form.html", product=product)

        db.session.add(product)
        db.session.commit()
        flash("商品创建成功", "success")
        return redirect(url_for("products.list_products"))

    return render_template("product_form.html", product=None)


@product_bp.route("/<int:product_id>/edit", methods=["GET", "POST"])
@login_required("admin")
def edit_product(product_id):
    product = db.session.get(Product, product_id)
    if product is None:
        flash("商品不存在", "danger")
        return redirect(url_for("products.list_products"))

    if request.method == "POST":
        form = request.form
        files = request.files

        product.model = form.get("model", "").strip()
        product.sku = form.get("sku", "").strip()
        product.package_size = form.get("package_size", "").strip()
        product.outer_size = form.get("outer_size", "").strip()
        product.inner_size = form.get("inner_size", "").strip()
        product.gross_weight = form.get("gross_weight", "").strip()
        product.net_weight = form.get("net_weight", "").strip()
        product.temperature_range = form.get("temperature_range", "").strip()
        product.power = form.get("power", "").strip()
        product.power_consumption = form.get("power_consumption", "").strip()
        product.compressor_brand = form.get("compressor_brand", "").strip()
        product.cooling_pipe_material = form.get(
            "cooling_pipe_material", ""
        ).strip()
        product.energy_level = form.get("energy_level", "").strip()

        save_uploaded_images(product, files)

        if not product.model:
            flash("型号为必填项", "danger")
            return render_template("product_form.html", product=product)

        db.session.commit()
        flash("商品更新成功", "success")
        return redirect(url_for("products.list_products"))

    return render_template("product_form.html", product=product)


@product_bp.route("/<int:product_id>/delete", methods=["POST"])
@login_required("admin")
def delete_product(product_id):
    product = db.session.get(Product, product_id)
    if product is None:
        flash("商品不存在", "danger")
        return redirect(url_for("products.list_products"))

    db.session.delete(product)
    db.session.commit()
    flash("商品已删除", "success")
    return redirect(url_for("products.list_products"))


def save_uploaded_images(product: Product, files):
    upload_folder = current_app.config["UPLOAD_FOLDER"]
    os.makedirs(upload_folder, exist_ok=True)

    real_file = files.get("image_real")
    param_file = files.get("image_param")

    if real_file and real_file.filename:
        filename = f"real_{product.id or 'new'}_{real_file.filename}"
        path = os.path.join(upload_folder, filename)
        real_file.save(path)
        product.image_real = filename

    if param_file and param_file.filename:
        filename = f"param_{product.id or 'new'}_{param_file.filename}"
        path = os.path.join(upload_folder, filename)
        param_file.save(path)
        product.image_param = filename


@product_bp.route("/import", methods=["GET", "POST"])
@login_required("admin")
def import_products():
    """从 Excel 批量导入商品。

    要求 Excel 存在以下列名（可以是中文表头，代码中对应映射可调整）：
    型号, SKU, 包装尺寸, 外形尺寸, 内部尺寸, 毛重, 净重, 温度范围, 功率, 耗电量, 压缩机品牌, 制冷管路材质, 能耗等级
    """
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash("请选择要上传的 Excel 文件", "danger")
            return render_template("product_import.html")

        import io

        from openpyxl import load_workbook

        try:
            in_memory = io.BytesIO(file.read())
            wb = load_workbook(in_memory, data_only=True)
            sheet = wb.active

            # 读取表头行
            headers = {}
            for idx, cell in enumerate(sheet[1], start=1):
                headers[cell.value] = idx

            def col(name: str):
                return headers.get(name)

            created = 0
            for row in sheet.iter_rows(min_row=2):
                model_cell = row[col("型号") - 1] if col("型号") else None
                if not model_cell or not model_cell.value:
                    continue

                product = Product(
                    model=str(model_cell.value).strip(),
                    sku=_get_cell(row, col("SKU")),
                    package_size=_get_cell(row, col("包装尺寸")),
                    outer_size=_get_cell(row, col("外形尺寸")),
                    inner_size=_get_cell(row, col("内部尺寸")),
                    gross_weight=_get_cell(row, col("毛重")),
                    net_weight=_get_cell(row, col("净重")),
                    temperature_range=_get_cell(row, col("温度范围")),
                    power=_get_cell(row, col("功率")),
                    power_consumption=_get_cell(row, col("耗电量")),
                    compressor_brand=_get_cell(row, col("压缩机品牌")),
                    cooling_pipe_material=_get_cell(row, col("制冷管路材质")),
                    energy_level=_get_cell(row, col("能耗等级")),
                )
                db.session.add(product)
                created += 1

            db.session.commit()
            flash(f"成功导入 {created} 条商品记录", "success")
            return redirect(url_for("products.list_products"))
        except Exception as exc:  # noqa: BLE001
            flash(f"导入失败: {exc}", "danger")

    return render_template("product_import.html")


def _get_cell(row, col_index):
    if not col_index:
        return ""
    cell = row[col_index - 1]
    return "" if cell.value is None else str(cell.value).strip()

